# -*- coding: utf-8 -*-
"""
Tự động tạo các file template Excel (.xlsx) chuẩn theo Thông tư 10/2024/TT-BTNMT.
"""

import os

def create_so_dia_chinh_template(output_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sổ Địa Chính"
    ws.views.sheetView[0].showGridLines = True
    
    # Fonts & Styles
    font_title = Font(name="Times New Roman", size=16, bold=True)
    font_subtitle = Font(name="Times New Roman", size=11, italic=True)
    font_header = Font(name="Times New Roman", size=10, bold=True)
    font_watermark = Font(name="Times New Roman", size=12, bold=True, color="FF0000")
    
    fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_side = Side(border_style="thin", color="000000")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # 1. Ghi Watermark "ẢNH MẪU / BẢN MẪU THỬ NGHIỆM"
    ws.merge_cells("A1:K1")
    ws["A1"] = "⚠️ ẢNH MẪU / BẢN MẪU THỬ NGHIỆM (THÔNG TƯ 10/2024/TT-BTNMT)"
    ws["A1"].font = font_watermark
    ws["A1"].alignment = align_center
    ws.row_dimensions[1].height = 25
    
    # 2. Tiêu đề chính
    ws.merge_cells("A2:K2")
    ws["A2"] = "Mẫu số 01/ĐK"
    ws["A2"].font = Font(name="Times New Roman", size=11, bold=True)
    ws["A2"].alignment = Alignment(horizontal="right")
    
    ws.merge_cells("A3:K3")
    ws["A3"] = "SỔ ĐỊA CHÍNH"
    ws["A3"].font = font_title
    ws["A3"].alignment = align_center
    ws.row_dimensions[3].height = 30
    
    ws.merge_cells("A4:K4")
    ws["A4"] = "Xã/Phường: ......................... Huyện/Quận: ......................... Tỉnh/Thành phố: ........................."
    ws["A4"].font = Font(name="Times New Roman", size=11, bold=True)
    ws["A4"].alignment = align_center
    
    # 3. Header bảng
    headers = [
        "STT", "Số tờ bản đồ", "Số thửa đất", "Tên người sử dụng đất / Chủ sở hữu",
        "Địa chỉ thửa đất", "Diện tích (m²)", "Hình thức sử dụng", "Mục đích sử dụng (Mã đất)",
        "Thời hạn sử dụng", "Nguồn gốc sử dụng đất", "Thông tin biến động / Ghi chú"
    ]
    
    ws.row_dimensions[6].height = 35
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
        
    # Thêm dòng số thứ tự cột (1, 2, 3...)
    ws.row_dimensions[7].height = 18
    for col_idx in range(1, 12):
        cell = ws.cell(row=7, column=col_idx, value=col_idx)
        cell.font = font_subtitle
        cell.alignment = align_center
        cell.border = border_all
        
    # Thiết lập độ rộng cột mẫu
    col_widths = {1: 6, 2: 12, 3: 12, 4: 25, 5: 25, 6: 15, 7: 15, 8: 15, 9: 15, 10: 25, 11: 30}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        
    # Lưu file
    wb.save(output_path)

def create_so_cap_gcn_template(output_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sổ Cấp GCN"
    ws.views.sheetView[0].showGridLines = True
    
    font_title = Font(name="Times New Roman", size=16, bold=True)
    font_subtitle = Font(name="Times New Roman", size=11, italic=True)
    font_header = Font(name="Times New Roman", size=10, bold=True)
    font_watermark = Font(name="Times New Roman", size=12, bold=True, color="FF0000")
    fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_all = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    
    # Watermark
    ws.merge_cells("A1:I1")
    ws["A1"] = "⚠️ ẢNH MẪU / BẢN MẪU THỬ NGHIỆM (THÔNG TƯ 10/2024/TT-BTNMT)"
    ws["A1"].font = font_watermark
    ws["A1"].alignment = align_center
    ws.row_dimensions[1].height = 25
    
    # Tiêu đề
    ws.merge_cells("A2:I2")
    ws["A2"] = "Mẫu số 02/ĐK"
    ws["A2"].font = Font(name="Times New Roman", size=11, bold=True)
    ws["A2"].alignment = Alignment(horizontal="right")
    
    ws.merge_cells("A3:I3")
    ws["A3"] = "SỔ CẤP GIẤY CHỨNG NHẬN QUYỀN SỬ DỤNG ĐẤT"
    ws["A3"].font = font_title
    ws["A3"].alignment = align_center
    ws.row_dimensions[3].height = 30
    
    headers = [
        "STT", "Tên người sử dụng đất / Chủ sở hữu", "Số tờ bản đồ", "Số thửa đất",
        "Diện tích (m²)", "Số phát hành GCN (Seri)", "Số vào sổ cấp GCN",
        "Ngày ký GCN", "Người nhận GCN (Ký nhận)"
    ]
    
    ws.row_dimensions[5].height = 35
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
        
    for col_idx in range(1, 10):
        cell = ws.cell(row=6, column=col_idx, value=col_idx)
        cell.font = font_subtitle
        cell.alignment = align_center
        cell.border = border_all
        
    col_widths = {1: 6, 2: 30, 3: 12, 4: 12, 5: 15, 6: 18, 7: 18, 8: 15, 9: 25}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        
    wb.save(output_path)

def create_so_muc_ke_template(output_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sổ Mục Kê"
    ws.views.sheetView[0].showGridLines = True
    
    font_title = Font(name="Times New Roman", size=16, bold=True)
    font_subtitle = Font(name="Times New Roman", size=11, italic=True)
    font_header = Font(name="Times New Roman", size=10, bold=True)
    font_watermark = Font(name="Times New Roman", size=12, bold=True, color="FF0000")
    fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_all = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    
    # Watermark
    ws.merge_cells("A1:H1")
    ws["A1"] = "⚠️ ẢNH MẪU / BẢN MẪU THỬ NGHIỆM (THÔNG TƯ 10/2024/TT-BTNMT)"
    ws["A1"].font = font_watermark
    ws["A1"].alignment = align_center
    ws.row_dimensions[1].height = 25
    
    # Tiêu đề
    ws.merge_cells("A2:H2")
    ws["A2"] = "Phụ lục số 15"
    ws["A2"].font = Font(name="Times New Roman", size=11, bold=True)
    ws["A2"].alignment = Alignment(horizontal="right")
    
    ws.merge_cells("A3:H3")
    ws["A3"] = "SỔ MỤC KÊ ĐẤT ĐAI"
    ws["A3"].font = font_title
    ws["A3"].alignment = align_center
    ws.row_dimensions[3].height = 30
    
    headers = [
        "STT", "Số tờ bản đồ", "Số thửa đất", "Tên người sử dụng / Quản lý đất",
        "Đối tượng sử dụng", "Diện tích (m²)", "Loại đất (Mã loại)", "Ghi chú"
    ]
    
    ws.row_dimensions[5].height = 35
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
        
    for col_idx in range(1, 9):
        cell = ws.cell(row=6, column=col_idx, value=col_idx)
        cell.font = font_subtitle
        cell.alignment = align_center
        cell.border = border_all
        
    col_widths = {1: 6, 2: 12, 3: 12, 4: 30, 5: 18, 6: 15, 7: 15, 8: 25}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        
    wb.save(output_path)

if __name__ == "__main__":
    templates_dir = os.path.dirname(__file__)
    os.makedirs(templates_dir, exist_ok=True)
    
    create_so_dia_chinh_template(os.path.join(templates_dir, "mau_01dk_so_dia_chinh.xlsx"))
    create_so_cap_gcn_template(os.path.join(templates_dir, "mau_02dk_so_cap_gcn.xlsx"))
    create_so_muc_ke_template(os.path.join(templates_dir, "so_muc_ke.xlsx"))
    print("Đã tạo xong 3 template mẫu biểu địa chính Excel!")
