# -*- coding: utf-8 -*-
"""
from ..common.common_utils import log_critical, log_warning

Module ghi dữ liệu thửa đất vào các mẫu biểu Excel địa chính (Excel Writer).
"""

from ..common.common_utils import log_warning

def write_cadastral_report(template_path, output_path, data_rows, report_type="so_dia_chinh", extra_info=None):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    """
    Ghi dữ liệu thửa đất vào file mẫu Excel.
    
    Args:
        template_path (str): Đường dẫn đến file template (.xlsx).
        output_path (str): Đường dẫn đến file kết quả đầu ra (.xlsx).
        data_rows (list): Danh sách các dict dữ liệu thửa đất, mỗi thửa dạng:
            {
                "sothua": str,
                "soto": str,
                "loaidat": str,
                "tenchu": str,
                "dientich": float,
                "diachi": str,
                "hinhthuc": str,
                "thoihan": str,
                "nguongoc": str,
                "ghichu": str
            }
        report_type (str): Loại báo cáo ("so_dia_chinh", "so_cap_gcn", "so_muc_ke").
        extra_info (dict): Các thông tin bổ sung hành chính (xa, huyen, tinh, ngay, nguoi_lap).
        
    Returns:
        bool: True nếu thành công, False nếu thất bại.
    """
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # 1. Điền thông tin hành chính tiêu đề nếu có
        info = extra_info or {}
        xa = info.get("xa", "")
        huyen = info.get("huyen", "")
        tinh = info.get("tinh", "")
        nguoi_lap = info.get("nguoi_lap", "")
        
        # Cập nhật địa chỉ xã/huyện/tỉnh ở tiêu đề
        if report_type == "so_dia_chinh" and ws["A4"].value:
            ws["A4"] = f"Xã/Phường: {xa}  Huyện/Quận: {huyen}  Tỉnh/Thành phố: {tinh}"
            
        # 2. Xác định vị trí ghi dữ liệu ròng đầu tiên (start_row)
        # Sổ địa chính: bắt đầu từ dòng 8 (dòng 6 là header, 7 là stt cột)
        # Sổ cấp GCN: bắt đầu từ dòng 7 (dòng 5 là header, 6 là stt cột)
        # Sổ mục kê: bắt đầu từ dòng 7 (dòng 5 là header, 6 là stt cột)
        if report_type == "so_dia_chinh":
            start_row = 8
        else:
            start_row = 7
            
        # Font và viền cho các ô dữ liệu mới chèn vào
        font_data = Font(name="Times New Roman", size=10)
        thin_side = Side(border_style="thin", color="000000")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
        
        current_row = start_row
        total_area = 0.0
        
        # 3. Ghi từng dòng dữ liệu
        for idx, data in enumerate(data_rows, 1):
            ws.row_dimensions[current_row].height = 24
            
            # Ghi các cột tùy thuộc vào loại báo cáo
            if report_type == "so_dia_chinh":
                # Mẫu 01/ĐK: STT, Số tờ, Số thửa, Tên chủ, Địa chỉ, Diện tích, Hình thức, Loại đất, Thời hạn, Nguồn gốc, Biến động
                row_data = [
                    idx,
                    data.get("soto", ""),
                    data.get("sothua", ""),
                    data.get("tenchu", ""),
                    data.get("diachi", ""),
                    data.get("dientich", 0.0),
                    data.get("hinhthuc", "Riêng"),
                    data.get("loaidat", "Khac"),
                    data.get("thoihan", "Lâu dài"),
                    data.get("nguongoc", "Được nhà nước giao đất"),
                    data.get("ghichu", "")
                ]
            elif report_type == "so_cap_gcn":
                # Mẫu 02/ĐK: STT, Tên chủ, Số tờ, Số thửa, Diện tích, Số Seri, Số vào sổ, Ngày ký, Người nhận GCN
                row_data = [
                    idx,
                    data.get("tenchu", ""),
                    data.get("soto", ""),
                    data.get("sothua", ""),
                    data.get("dientich", 0.0),
                    data.get("seri", ""),
                    data.get("so_vao_so", ""),
                    data.get("ngay_ky", ""),
                    data.get("nguoi_nhan", "")
                ]
            else: # so_muc_ke
                # Sổ mục kê: STT, Số tờ, Số thửa, Tên chủ, Đối tượng sử dụng, Diện tích, Loại đất, Ghi chú
                row_data = [
                    idx,
                    data.get("soto", ""),
                    data.get("sothua", ""),
                    data.get("tenchu", ""),
                    data.get("doi_tuong", "Gia đình/Cá nhân"),
                    data.get("dientich", 0.0),
                    data.get("loaidat", "Khac"),
                    data.get("ghichu", "")
                ]
                
            # Duyệt qua từng cột để ghi vào Excel
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = font_data
                cell.border = border_all
                
                # Căn lề phù hợp cho từng loại cột
                if col_idx in [1, 2, 3, 7, 8, 9] if report_type == "so_dia_chinh" else [1, 3, 4, 6, 7, 8]:
                    cell.alignment = align_center
                elif isinstance(val, (int, float)):
                    cell.alignment = align_right
                    cell.number_format = "#,##0.0" # Định dạng số thập phân đẹp
                else:
                    cell.alignment = align_left
                    
            # Cộng dồn tổng diện tích
            try:
                total_area += float(data.get("dientich", 0.0))
            except Exception as e:  # noqa: BLE001
                import traceback
                log_warning(f"[write_cadastral_report loop] Lỗi bị bỏ qua: {e}\n{traceback.format_exc()}")
                
            current_row += 1
            
        # 4. Ghi dòng Tổng cộng (Summary Row)
        ws.row_dimensions[current_row].height = 26
        font_bold = Font(name="Times New Roman", size=10, bold=True)
        
        # Merge các cột đầu để ghi chữ "Tổng cộng"
        merge_limit = 5 if report_type in ("so_dia_chinh", "so_muc_ke") else 4
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=merge_limit)
        
        cell_total_lbl = ws.cell(row=current_row, column=1, value=f"TỔNG CỘNG ({len(data_rows)} thửa)")
        cell_total_lbl.font = font_bold
        cell_total_lbl.alignment = align_center
        
        # Điền viền cho các ô đã merge
        for col_idx in range(1, merge_limit + 1):
            ws.cell(row=current_row, column=col_idx).border = border_all
            
        # Ghi giá trị tổng diện tích vào cột diện tích tương ứng
        area_col = 6 if report_type in ("so_dia_chinh", "so_muc_ke") else 5
        cell_total_area = ws.cell(row=current_row, column=area_col, value=total_area)
        cell_total_area.font = font_bold
        cell_total_area.border = border_all
        cell_total_area.alignment = align_right
        cell_total_area.number_format = "#,##0.0"
        
        # Viền cho các cột trống còn lại của dòng tổng cộng
        num_cols = 11 if report_type == "so_dia_chinh" else (9 if report_type == "so_cap_gcn" else 8)
        for col_idx in range(area_col + 1, num_cols + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = border_all
            
        # 5. Ghi phần ký tên đóng dấu ở cuối báo cáo
        current_row += 2
        ws.row_dimensions[current_row].height = 20
        ws.cell(row=current_row, column=2, value="Người lập biểu").font = font_bold
        ws.cell(row=current_row, column=2).alignment = align_center
        
        right_col = 10 if report_type == "so_dia_chinh" else (8 if report_type == "so_cap_gcn" else 7)
        ws.cell(row=current_row, column=right_col, value="Ngày .... tháng .... năm 20...").font = Font(name="Times New Roman", size=10, italic=True)
        ws.cell(row=current_row, column=right_col).alignment = align_center
        
        current_row += 1
        ws.row_dimensions[current_row].height = 20
        ws.cell(row=current_row, column=2, value="(Ký, ghi rõ họ tên)").font = Font(name="Times New Roman", size=9, italic=True)
        ws.cell(row=current_row, column=2).alignment = align_center
        
        ws.cell(row=current_row, column=right_col, value="ỦY BAN NHÂN DÂN XÃ/PHƯỜNG").font = font_bold
        ws.cell(row=current_row, column=right_col).alignment = align_center
        
        current_row += 1
        ws.row_dimensions[current_row].height = 20
        ws.cell(row=current_row, column=right_col, value="(Ký tên, đóng dấu)").font = Font(name="Times New Roman", size=9, italic=True)
        ws.cell(row=current_row, column=right_col).alignment = align_center
        
        # Điền họ tên người lập nếu có nhập
        if nguoi_lap:
            current_row += 4
            ws.cell(row=current_row, column=2, value=nguoi_lap).font = font_bold
            ws.cell(row=current_row, column=2).alignment = align_center
            
        wb.save(output_path)
        return True
    except PermissionError:
        from qgis.PyQt.QtWidgets import QMessageBox
        QMessageBox.critical(
            None,
            "Không thể lưu file",
            "File Excel đang được mở bởi chương trình khác.\n"
            "Vui lòng đóng file rồi thử lại."
        )
        return False
    except Exception as e:  # noqa: BLE001
        import traceback
        log_warning(f"Lỗi khi ghi báo cáo Excel: {e}\n{traceback.format_exc()}")
        return False
