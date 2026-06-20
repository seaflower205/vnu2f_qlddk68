# -*- coding: utf-8 -*-
import os
import csv
from qgis.PyQt.QtWidgets import QFileDialog, QMessageBox
from qgis.core import Qgis
from qgis.utils import iface

class StatsExporter:
    @staticmethod
    def export_csv(stats_data, parent_widget):
        """Xuất dữ liệu thống kê bảng ra file CSV."""
        if not stats_data:
            QMessageBox.warning(parent_widget, "Cảnh báo", "Không có dữ liệu thống kê để xuất.")
            return

        file_path, _ = QFileDialog.getSaveFileName(parent_widget, "Xuất thống kê CSV", "", "CSV Files (*.csv)")
        if not file_path:
            return

        try:
            with open(file_path, mode="w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Mã loại đất", "Tên loại đất", "Số lượng thửa", "Tổng diện tích (m2)", "Tổng diện tích (ha)", "Tỷ lệ (%)"])
                
                total_cnt = 0
                total_area = 0.0
                for data in stats_data:
                    writer.writerow([
                        data["code"], data["name_vi"], data["count"], 
                        data["area_m2"], data["area_ha"], f"{data['percentage']:.2f}"
                    ])
                    total_cnt += data["count"]
                    total_area += data["area_m2"]
                    
                writer.writerow(["TỔNG CỘNG", "", total_cnt, total_area, total_area / 10000.0, "100.00"])

            if iface:
                iface.messageBar().pushMessage(
                    "Thành công", f"Đã xuất file báo cáo thống kê CSV:\n{os.path.basename(file_path)}",
                    level=Qgis.Success, duration=5
                )
        except Exception as e:  # noqa: BLE001
            QMessageBox.critical(parent_widget, "Lỗi", f"Không thể xuất file CSV: {str(e)}")

    @staticmethod
    def export_excel(stats_data, parent_widget):
        """Xuất dữ liệu thống kê bảng ra file Excel dùng openpyxl."""
        if not stats_data:
            QMessageBox.warning(parent_widget, "Cảnh báo", "Không có dữ liệu thống kê để xuất.")
            return

        try:
            import openpyxl
            from openpyxl.styles import (
                Alignment,
                Border,
                Font as XlsFont,
                PatternFill,
                Side,
            )
        except ImportError:
            QMessageBox.warning(
                parent_widget, "Thiếu thư viện", 
                "Thư viện 'openpyxl' không có sẵn. Vui lòng cài đặt qua Python hoặc dùng tính năng Xuất CSV."
            )
            return

        file_path, _ = QFileDialog.getSaveFileName(parent_widget, "Xuất báo cáo thống kê Excel", "", "Excel Files (*.xlsx)")
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Thống kê Đất đai"

            font_title = XlsFont(name="Arial", size=16, bold=True)
            font_header = XlsFont(name="Arial", size=11, bold=True, color="FFFFFF")
            font_data = XlsFont(name="Arial", size=10)
            font_total = XlsFont(name="Arial", size=10, bold=True)
            
            fill_header = PatternFill(start_color="18181B", end_color="18181B", fill_type="solid")
            fill_total = PatternFill(start_color="F4F4F5", end_color="F4F4F5", fill_type="solid")
            
            border_thin = Border(
                left=Side(style="thin", color="D4D4D8"),
                right=Side(style="thin", color="D4D4D8"),
                top=Side(style="thin", color="D4D4D8"),
                bottom=Side(style="thin", color="D4D4D8")
            )
            border_total = Border(
                top=Side(style="thin", color="18181B"),
                bottom=Side(style="double", color="18181B")
            )

            ws.merge_cells("A1:F1")
            ws["A1"] = "BẢNG TỔNG HỢP DIỆN TÍCH CƠ CẤU ĐẤT ĐAI"
            ws["A1"].font = font_title
            ws["A1"].alignment = Alignment(horizontal="center")
            ws.row_dimensions[1].height = 30

            ws.append([])

            headers = ["Mã loại đất", "Tên loại đất", "Số lượng thửa", "Tổng diện tích (m²)", "Tổng diện tích (ha)", "Tỷ lệ (%)"]
            ws.append(headers)
            ws.row_dimensions[3].height = 24
            
            for col_idx in range(1, 7):
                cell = ws.cell(row=3, column=col_idx)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border_thin

            total_cnt = 0
            total_area = 0.0
            
            for data in stats_data:
                row_data = [
                    data["code"], data["name_vi"], data["count"], 
                    data["area_m2"], data["area_ha"], data["percentage"] / 100.0
                ]
                ws.append(row_data)
                
                curr_row = ws.max_row
                ws.row_dimensions[curr_row].height = 20
                
                for col_idx in range(1, 7):
                    cell = ws.cell(row=curr_row, column=col_idx)
                    cell.font = font_data
                    cell.border = border_thin
                    
                    if col_idx in [1, 2]:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif col_idx == 3:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = "#,##0"
                    elif col_idx == 4:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = "#,##0.0"
                    elif col_idx == 5:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = "#,##0.0000"
                    elif col_idx == 6:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = "0.00%"
                        
                total_cnt += data["count"]
                total_area += data["area_m2"]

            tot_row = ws.max_row + 1
            ws.cell(row=tot_row, column=1, value="TỔNG CỘNG").font = font_total
            ws.cell(row=tot_row, column=2, value="")
            ws.cell(row=tot_row, column=3, value=total_cnt).font = font_total
            ws.cell(row=tot_row, column=4, value=total_area).font = font_total
            ws.cell(row=tot_row, column=5, value=total_area / 10000.0).font = font_total
            ws.cell(row=tot_row, column=6, value=1.0).font = font_total
            
            ws.row_dimensions[tot_row].height = 22
            
            for col_idx in range(1, 7):
                cell = ws.cell(row=tot_row, column=col_idx)
                cell.fill = fill_total
                cell.border = border_total
                
                if col_idx in [1, 2]:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif col_idx == 3:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "#,##0"
                elif col_idx == 4:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "#,##0.0"
                elif col_idx == 5:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "#,##0.0000"
                elif col_idx == 6:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "0.00%"

            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            wb.save(file_path)
            
            if iface:
                iface.messageBar().pushMessage(
                    "Thành công", f"Đã xuất báo cáo thống kê Excel thành công:\n{os.path.basename(file_path)}",
                    level=Qgis.Success, duration=5
                )
        except Exception as e:  # noqa: BLE001
            QMessageBox.critical(parent_widget, "Lỗi", f"Không thể xuất file Excel: {str(e)}")
